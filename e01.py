# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook



file_path = 'c:/aat/data/pets.xlsx'

# Load in the workbook
wb = load_workbook( file_path )

# Get sheet names
print(wb.get_sheet_names())

# Get a sheet by name
sheet = wb.get_sheet_by_name( 'Sheet1' )

# Print the sheet title
print( sheet.title )

col = 'A'
i    = 1

while True:

    try:
        c1 = 'A' + str( i ).strip()
        c2 = 'B' + str( i ).strip()


        c  = sheet[ col ]
        v1 = sheet[ c1 ].value
        v2 = sheet[ c2 ].value

        if v1 == None and v2 == None:
            break

        print( ' {}\t\t{} '.format( v1, v2 ) )

        i = i + 1

    except:
        break

    if i >= 10:
        break


print( '\n\n\nend ' )
