# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

def to_csv( xlsx_path, csv_path, sheet_name ):
    #

    # Load in the workbook
    wb = load_workbook( xlsx_path )

    # Get a sheet by name
    sheet = wb.get_sheet_by_name( sheet_name )

    col   = 'A'
    i     = 1
    with open( csv_path, 'w' ) as f:

        while True:
            try:
                c1 = 'A' + str( i ).strip()
                c2 = 'B' + str( i ).strip()
                c  = sheet[ col ]
                v1 = sheet[ c1 ].value
                v2 = sheet[ c2 ].value
                i  = i + 1

                if v1 == None and v2 == None:
                    break

                print( ' {}\t\t{} '.format( v1, v2 ) )
                f.write( '{}\t{}\n'.format(v1, v2) )

            except:
                break


    print( '\n\n\nend ' )



xlsx_path  = 'c:/aat/data/pets.xlsx'
csv_path   = 'c:/aat/data/pets.csv'
sheet_name = 'Sheet1'

to_csv( xlsx_path, csv_path, sheet_name )